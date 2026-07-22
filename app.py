import streamlit as st
import pydicom
import pydicom.sequence
import pydicom.multival
from pydicom.datadict import keyword_for_tag
from pydicom._dicom_dict import DicomDictionary
import pandas as pd
import io
import re
import base64
import zipfile
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="DICOM Tag Validator | SwiftMR",
    page_icon="🏥",
    layout="wide"
)

# ── Logo ─────────────────────────────────────────────
def get_image_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return None

logo_b64          = get_image_base64("SwiftMR Logo.png")
logo_html         = (f'<img src="data:image/png;base64,{logo_b64}" style="width:44px;height:44px;object-fit:contain;">' if logo_b64 else "🏥")
sidebar_logo_html = (f'<img src="data:image/png;base64,{logo_b64}" style="width:48px;height:48px;object-fit:contain;">' if logo_b64 else "🏥")

# ── CSS ──────────────────────────────────────────────
st.markdown("""
<style>
@media (prefers-color-scheme: dark) {
    .stApp { background-color: #0f1117 !important; }
    .airs-header { background: linear-gradient(135deg,#1a1f2e 0%,#0d1117 100%) !important; border-bottom: 2px solid #00d4ff !important; }
    .airs-title p { color: #8892a4 !important; }
    .airs-badge { background: rgba(0,212,255,0.1) !important; border: 1px solid rgba(0,212,255,0.3) !important; color: #00d4ff !important; }
    .summary-card  { background: linear-gradient(135deg,#1a1f2e,#141820) !important; border: 1px solid #2a3040 !important; }
    .section-card  { background: linear-gradient(135deg,#1a1f2e,#141820) !important; border: 1px solid #2a3040 !important; }
    .metric-card   { background: #1a1f2e !important; border: 1px solid #2a3040 !important; }
    .metric-value  { color: #e8eaf0 !important; }
    .metric-label  { color: #8892a4 !important; }
    .sidebar-section-title { color: #00d4ff !important; border-bottom: 1px solid #2a3040 !important; }
    .manufacturer-badge { background: rgba(0,212,255,0.1) !important; border: 1px solid rgba(0,212,255,0.3) !important; color: #00d4ff !important; }
    .file-problem-card { background: rgba(255,60,60,0.08) !important; border: 1px solid rgba(255,60,60,0.3) !important; }
    .phi-notice { background: linear-gradient(135deg,rgba(0,180,100,0.10),rgba(0,120,80,0.08)) !important; border: 1.5px solid rgba(0,200,120,0.4) !important; border-left: 4px solid #00c878 !important; }
    .cat1-header { background: linear-gradient(135deg,rgba(255,60,60,0.15),rgba(200,0,0,0.08)) !important; border: 1px solid rgba(255,60,60,0.3) !important; }
    .cat2-header { background: linear-gradient(135deg,rgba(255,140,0,0.15),rgba(200,100,0,0.08)) !important; border: 1px solid rgba(255,140,0,0.3) !important; }
    .cat3-header { background: linear-gradient(135deg,rgba(0,180,255,0.12),rgba(0,100,200,0.08)) !important; border: 1px solid rgba(0,180,255,0.3) !important; }
    .no-mfr-box  { background: rgba(255,180,0,0.08) !important; border: 1px solid rgba(255,180,0,0.3) !important; }
    .std-info-box { background: rgba(0,212,255,0.05) !important; border: 1px solid rgba(0,212,255,0.15) !important; }
}
@media (prefers-color-scheme: light) {
    .stApp { background-color: #f0f4f8 !important; }
    .airs-header { background: linear-gradient(135deg,#ffffff 0%,#e8f0fe 100%) !important; border-bottom: 2px solid #0066ff !important; }
    .airs-title p { color: #5a6a7a !important; }
    .airs-badge { background: rgba(0,102,255,0.1) !important; border: 1px solid rgba(0,102,255,0.3) !important; color: #0066ff !important; }
    .summary-card  { background: linear-gradient(135deg,#ffffff,#f5f8ff) !important; border: 1px solid #d0d8e8 !important; box-shadow: 0 4px 24px rgba(0,0,0,0.08) !important; }
    .section-card  { background: linear-gradient(135deg,#ffffff,#f5f8ff) !important; border: 1px solid #d0d8e8 !important; box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important; }
    .metric-card   { background: #ffffff !important; border: 1px solid #d0d8e8 !important; box-shadow: 0 2px 8px rgba(0,0,0,0.06) !important; }
    .metric-value  { color: #1a2030 !important; }
    .metric-label  { color: #5a6a7a !important; }
    .sidebar-section-title { color: #0066ff !important; border-bottom: 1px solid #d0d8e8 !important; }
    .manufacturer-badge { background: rgba(0,102,255,0.1) !important; border: 1px solid rgba(0,102,255,0.3) !important; color: #0066ff !important; }
    .file-problem-card { background: rgba(255,60,60,0.06) !important; border: 1px solid rgba(255,60,60,0.3) !important; }
    .phi-notice { background: linear-gradient(135deg,rgba(0,180,100,0.08),rgba(0,120,80,0.05)) !important; border: 1.5px solid rgba(0,180,100,0.4) !important; border-left: 4px solid #00a86b !important; }
    .cat1-header { background: linear-gradient(135deg,rgba(255,60,60,0.10),rgba(200,0,0,0.05)) !important; border: 1px solid rgba(255,60,60,0.3) !important; }
    .cat2-header { background: linear-gradient(135deg,rgba(255,140,0,0.10),rgba(200,100,0,0.05)) !important; border: 1px solid rgba(255,140,0,0.3) !important; }
    .cat3-header { background: linear-gradient(135deg,rgba(0,180,255,0.08),rgba(0,100,200,0.05)) !important; border: 1px solid rgba(0,180,255,0.3) !important; }
    .no-mfr-box  { background: rgba(255,180,0,0.06) !important; border: 1px solid rgba(255,180,0,0.3) !important; }
    .std-info-box { background: rgba(0,100,200,0.04) !important; border: 1px solid rgba(0,100,200,0.12) !important; }
}
.airs-header { display:flex; align-items:center; gap:16px; padding:20px 28px; margin-bottom:24px; border-radius:0 0 16px 16px; }
.airs-logo-box { width:52px; height:52px; background:transparent; border-radius:12px; display:flex; align-items:center; justify-content:center; flex-shrink:0; }
.airs-title h1 { margin:0; font-size:22px; font-weight:800; background:linear-gradient(90deg,#00d4ff,#0066ff); -webkit-background-clip:text; -webkit-text-fill-color:transparent; letter-spacing:2px; }
.airs-title p  { margin:2px 0 0; font-size:13px; letter-spacing:1px; }
.airs-badge    { margin-left:auto; padding:6px 14px; border-radius:20px; font-size:12px; font-weight:600; letter-spacing:1px; }
.summary-card  { border-radius:16px; padding:24px 28px; margin-bottom:20px; }
.section-card  { border-radius:16px; padding:20px 24px; margin-bottom:16px; }
.section-title { font-size:16px; font-weight:700; display:flex; align-items:center; gap:10px; margin-bottom:16px; }
.metric-card   { border-radius:12px; padding:16px 20px; text-align:center; }
.metric-value  { font-size:28px; font-weight:800; line-height:1.1; }
.metric-label  { font-size:11px; font-weight:600; letter-spacing:1px; text-transform:uppercase; margin-top:4px; }
.overall-pass    { background:linear-gradient(135deg,rgba(0,200,100,0.15),rgba(0,150,80,0.1)) !important; border:2px solid rgba(0,200,100,0.4) !important; }
.overall-fail    { background:linear-gradient(135deg,rgba(255,60,60,0.15),rgba(200,0,0,0.1)) !important; border:2px solid rgba(255,60,60,0.4) !important; }
.overall-warning { background:linear-gradient(135deg,rgba(255,180,0,0.15),rgba(200,140,0,0.1)) !important; border:2px solid rgba(255,180,0,0.4) !important; }
.overall-title { font-size:28px; font-weight:900; letter-spacing:2px; margin-bottom:4px; }
.overall-sub   { font-size:14px; opacity:0.8; }
.manufacturer-badge { display:inline-flex; align-items:center; gap:8px; padding:8px 16px; border-radius:20px; font-size:13px; font-weight:700; margin:4px; }
.sidebar-section-title { font-size:12px; font-weight:700; letter-spacing:1px; text-transform:uppercase; margin-bottom:10px; padding-bottom:6px; }
.phi-notice        { border-radius:12px; padding:16px 20px; margin-bottom:20px; }
.file-problem-card { border-radius:12px; padding:12px 16px; margin-bottom:8px; }
.cat1-header { border-radius:12px; padding:14px 20px; margin-bottom:12px; }
.cat2-header { border-radius:12px; padding:14px 20px; margin-bottom:12px; }
.cat3-header { border-radius:12px; padding:14px 20px; margin-bottom:12px; }
.no-mfr-box  { border-radius:12px; padding:16px 20px; margin-bottom:12px; }
.std-info-box { border-radius:10px; padding:12px 16px; margin-bottom:12px; font-size:12px; }
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="airs-header">
    <div class="airs-logo-box">{logo_html}</div>
    <div class="airs-title">
        <h1>SwiftMR</h1>
        <p>DICOM Tag Validator &nbsp;·&nbsp; Internal Tool</p>
    </div>
    <div class="airs-badge">v2.0</div>
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════
# DICOM Standard 기반 검증 엔진
# ════════════════════════════════════════════════════

# ── VR 검증 규칙 (DICOM PS3.5 Table 6.2-1) ──────────
VR_RULES = {
    'DS': {'cast': float, 'max_len': 16,  'desc': 'Decimal String'},
    'IS': {'cast': int,   'max_len': 12,  'range': (-2**31, 2**31-1), 'desc': 'Integer String'},
    'US': {'cast': int,   'range': (0, 65535),      'desc': 'Unsigned Short'},
    'SS': {'cast': int,   'range': (-32768, 32767),  'desc': 'Signed Short'},
    'UL': {'cast': int,   'range': (0, 2**32-1),    'desc': 'Unsigned Long'},
    'SL': {'cast': int,   'range': (-2**31, 2**31-1),'desc': 'Signed Long'},
    'FL': {'cast': float, 'desc': 'Floating Point Single'},
    'FD': {'cast': float, 'desc': 'Floating Point Double'},
    'CS': {'cast': str,   'max_len': 16,  'desc': 'Code String'},
    'LO': {'cast': str,   'max_len': 64,  'desc': 'Long String'},
    'SH': {'cast': str,   'max_len': 16,  'desc': 'Short String'},
    'ST': {'cast': str,   'max_len': 1024,'desc': 'Short Text'},
    'LT': {'cast': str,   'max_len': 10240,'desc': 'Long Text'},
    'UT': {'cast': str,   'desc': 'Unlimited Text'},
    'UI': {'cast': str,   'max_len': 64,  'pattern': r'^[0-9.]+$', 'desc': 'Unique Identifier'},
    'DA': {'cast': str,   'pattern': r'^\d{8}$', 'desc': 'Date YYYYMMDD'},
    'TM': {'cast': str,   'pattern': r'^\d{2}(\d{2}(\d{2}(\.\d+)?)?)?$', 'desc': 'Time'},
    'PN': {'cast': str,   'max_len': 324, 'desc': 'Person Name'},
    'AS': {'cast': str,   'pattern': r'^\d{3}[DWMY]$', 'desc': 'Age String'},
}

# ── DICOM Enum 값 (PS3.3 기반) ───────────────────────
DICOM_ENUMS = {
    (0x0008,0x0060): {  # Modality
        'values': {
            'AR','AU','BDUS','BI','BMD','CR','CT','DG','DOC','DX','ECG','EPS',
            'ES','FID','GM','HC','HD','IO','IOL','IVOCT','IVUS','KER','KO','LEN',
            'LS','MG','MR','M3D','NM','OAM','OCT','OP','OPM','OPT','OPTBSV',
            'OPTENF','OPV','OSS','OT','PLAN','PR','PT','PX','REG','RESP','RF',
            'RG','RTDOSE','RTIMAGE','RTINTENT','RTPLAN','RTRAD','RTSTRUCT','RWV',
            'SEG','SM','SMR','SR','SRF','STAIN','TEXTUREMAP','TG','US','VA','XA','XC'
        },
        'desc': 'Modality (PS3.3 C.7.3.1.1.1)'
    },
    (0x0018,0x5100): {  # PatientPosition
        'values': {
            'HFP','HFS','HFDR','HFDL',
            'FFP','FFS','FFDR','FFDL',
            'LFP','LFS','RFP','RFS',
            'AFDR','AFDL','PFDR','PFDL'
        },
        'desc': 'Patient Position (PS3.3 C.7.3.1.1.2)'
    },
    (0x0008,0x0008): {  # ImageType - Value 1만 검증
        'values': {'ORIGINAL','DERIVED'},
        'desc': 'Image Type Value1 (PS3.3 C.7.6.1.1.2)',
        'check_index': 0
    },
    (0x0018,0x0023): {  # MRAcquisitionType
        'values': {'2D','3D'},
        'desc': 'MR Acquisition Type (PS3.3 C.8.3.1.1.1)'
    },
    (0x0028,0x0004): {  # PhotometricInterpretation
        'values': {
            'MONOCHROME1','MONOCHROME2','PALETTE COLOR','RGB',
            'YBR_FULL','YBR_FULL_422','YBR_PARTIAL_422',
            'YBR_PARTIAL_420','YBR_ICT','YBR_RCT'
        },
        'desc': 'Photometric Interpretation (PS3.3 C.7.6.3.1.2)'
    },
    (0x0028,0x0103): {  # PixelRepresentation
        'values': {0, 1},
        'desc': 'Pixel Representation (PS3.3 C.7.6.3.1.5): 0=unsigned, 1=signed'
    },
    (0x0018,0x1312): {  # InPlanePhaseEncodingDirection
        'values': {'ROW','COL','OTHER'},
        'desc': 'In-Plane Phase Encoding Direction (PS3.3 C.8.3.1.1.1)'
    },
    (0x0018,0x0020): {  # ScanningSequence
        'values': {'SE','IR','GR','EPI','RM'},
        'desc': 'Scanning Sequence (PS3.3 C.8.3.1.1.4)'
    },
    (0x0018,0x0021): {  # SequenceVariant
        'values': {'SK','MTC','SS','TRSS','SP','MP','OSP','NONE'},
        'desc': 'Sequence Variant (PS3.3 C.8.3.1.1.5)'
    },
}

# ── 태그별 특수 규칙 (PS3.3 참조) ───────────────────
SPECIAL_RULES = {
    # ImagePositionPatient: VM=3, 전부 0이면 placeholder 의심
    (0x0020,0x0032): {
        'ref': 'PS3.3 C.7.6.2.1.1',
        'check': lambda vals: (
            (False, f"Expected 3 values (X\\Y\\Z), got {len(vals)}")
            if len(vals) != 3
            else (False, "All zeros — likely unset placeholder")
            if all(_is_zero(v) for v in vals)
            else (True, "")
        )
    },
    # ImageOrientationPatient: VM=6, 전부 0이면 placeholder 의심
    (0x0020,0x0037): {
        'ref': 'PS3.3 C.7.6.2.1.1',
        'check': lambda vals: (
            (False, f"Expected 6 values, got {len(vals)}")
            if len(vals) != 6
            else (False, "All zeros — likely unset placeholder")
            if all(_is_zero(v) for v in vals)
            else (True, "")
        )
    },
    # PixelSpacing: VM=2, 양수
    (0x0028,0x0030): {
        'ref': 'PS3.3 10.7.1.3',
        'check': lambda vals: (
            (False, f"Expected 2 values (row\\col spacing), got {len(vals)}")
            if len(vals) != 2
            else (False, f"PixelSpacing must be positive, got {[str(v) for v in vals]}")
            if any(_to_float(v) is not None and _to_float(v) <= 0 for v in vals)
            else (True, "")
        )
    },
    # WindowWidth: 반드시 양수 (PS3.3 C.7.6.3.1.5)
    (0x0028,0x1051): {
        'ref': 'PS3.3 C.7.6.3.1.5',
        'check': lambda vals: (
            (False, f"WindowWidth must be > 0, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # BitsStored: 1~16 (PS3.5 8.2)
    (0x0028,0x0101): {
        'ref': 'PS3.5 8.2',
        'check': lambda vals: (
            (False, f"BitsStored must be 1-16, got {vals[0]}")
            if vals and not (1 <= int(str(vals[0])) <= 16)
            else (True, "")
        )
    },
    # SliceThickness: 음수 불가 (Type2 → 0 허용)
    (0x0018,0x0050): {
        'ref': 'PS3.3 C.7.6.2.1.1',
        'check': lambda vals: (
            (False, f"SliceThickness cannot be negative, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) < 0
            else (True, "")
        )
    },
    # SpacingBetweenSlices: 양수
    (0x0018,0x0088): {
        'ref': 'PS3.3 C.7.6.3.1.4',
        'check': lambda vals: (
            (False, f"SpacingBetweenSlices must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # Rows / Columns: 양수
    (0x0028,0x0010): {
        'ref': 'PS3.3 C.7.6.3.1.4',
        'check': lambda vals: (
            (False, f"Rows must be positive, got {vals[0]}")
            if vals and int(str(vals[0])) <= 0
            else (True, "")
        )
    },
    (0x0028,0x0011): {
        'ref': 'PS3.3 C.7.6.3.1.4',
        'check': lambda vals: (
            (False, f"Columns must be positive, got {vals[0]}")
            if vals and int(str(vals[0])) <= 0
            else (True, "")
        )
    },
    # ImagingFrequency: 양수
    (0x0018,0x0084): {
        'ref': 'PS3.3 C.8.3.1.1.1',
        'check': lambda vals: (
            (False, f"ImagingFrequency must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # MagneticFieldStrength: 양수
    (0x0018,0x0087): {
        'ref': 'PS3.3 C.8.3.1.1.1',
        'check': lambda vals: (
            (False, f"MagneticFieldStrength must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # NumberOfAverages: 양수
    (0x0018,0x0083): {
        'ref': 'PS3.3 C.8.3.1.1.1',
        'check': lambda vals: (
            (False, f"NumberOfAverages must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # EchoTime: 양수
    (0x0018,0x0081): {
        'ref': 'PS3.3 C.8.3.1.1.1',
        'check': lambda vals: (
            (False, f"EchoTime must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
    # RepetitionTime: 양수
    (0x0018,0x0080): {
        'ref': 'PS3.3 C.8.3.1.1.1',
        'check': lambda vals: (
            (False, f"RepetitionTime must be positive, got {vals[0]}")
            if vals and _to_float(vals[0]) is not None and _to_float(vals[0]) <= 0
            else (True, "")
        )
    },
}


def _is_zero(v):
    f = _to_float(v)
    return f is not None and f == 0.0

def _to_float(v):
    try:
        return float(str(v).strip())
    except Exception:
        return None

def _to_values(val):
    """pydicom 값 → 리스트"""
    if isinstance(val, pydicom.multival.MultiValue):
        return list(val)
    if isinstance(val, (list, tuple)):
        return list(val)
    return [val]

def parse_vm(vm_str):
    """
    VM 문자열 파싱
    '3' → (3,3), '1-n' → (1,None), '2-2n' → (2,None,2), '1-3' → (1,3)
    """
    vm_str = str(vm_str).strip()
    if '-' in vm_str:
        parts   = vm_str.split('-')
        min_vm  = int(parts[0])
        max_part = parts[1]
        if max_part in ('n',):
            return min_vm, None, 1
        elif max_part.endswith('n'):
            factor = int(max_part[:-1]) if max_part[:-1].isdigit() else 1
            return min_vm, None, factor
        else:
            return min_vm, int(max_part), 1
    else:
        n = int(vm_str)
        return n, n, 1

def get_std_info(tag_tuple):
    """DicomDictionary에서 표준 VR/VM/Name 조회"""
    tag_int = (tag_tuple[0] << 16) | tag_tuple[1]
    info    = DicomDictionary.get(tag_int)
    if info:
        return {'vr': info[0], 'vm': info[1], 'name': info[2],
                'retired': info[3], 'keyword': info[4]}
    return None

# ════════════════════════════════════════════════════
# 핵심 검증 함수 (5단계)
# ════════════════════════════════════════════════════
def validate_value_full(tag_tuple, elem):
    """
    5단계 DICOM 표준 기반 검증
    반환: (is_valid, issue_msg, std_info_dict)
    """
    if elem is None:
        return True, "", {}

    std = get_std_info(tag_tuple)

    try:
        val = elem.value

        # 빈값 체크
        if val is None:
            return False, "Null value", std or {}
        if isinstance(val, str) and val.strip() == "":
            return False, "Empty string value", std or {}
        if isinstance(val, pydicom.sequence.Sequence) and len(val) == 0:
            return False, "Empty sequence", std or {}

        # Binary/Sequence → 내용 검증 생략
        if isinstance(val, (bytes, bytearray)):
            return True, "", std or {}
        if isinstance(val, pydicom.sequence.Sequence):
            return True, "", std or {}

        values = _to_values(val)
        std_vr = std['vr'] if std else elem.VR
        std_vm = std['vm'] if std else '1-n'

        # ── Step 1: VM 검증 ──────────────────────
        if std_vm and std_vm not in ('', '1-n', '1'):
            try:
                min_vm, max_vm, factor = parse_vm(std_vm)
                count = len(values)
                if count < min_vm:
                    return False, f"VM violation: expected min {min_vm} values (VM={std_vm}), got {count}", std or {}
                if max_vm is not None and count > max_vm:
                    return False, f"VM violation: expected max {max_vm} values (VM={std_vm}), got {count}", std or {}
            except Exception:
                pass

        # ── Step 2: VR 타입 검증 ─────────────────
        rule = VR_RULES.get(std_vr)
        if rule:
            for v in values:
                s = str(v).strip()
                # 타입 변환
                try:
                    converted = rule['cast'](s)
                except (ValueError, TypeError):
                    return False, f"VR={std_vr} ({rule['desc']}): cannot convert '{s}' to {rule['cast'].__name__}", std or {}
                # 범위
                if 'range' in rule:
                    lo, hi = rule['range']
                    if not (lo <= converted <= hi):
                        return False, f"VR={std_vr}: value {converted} out of range [{lo}, {hi}]", std or {}
                # 패턴
                if 'pattern' in rule:
                    if not re.match(rule['pattern'], s):
                        return False, f"VR={std_vr} ({rule['desc']}): '{s}' does not match expected format", std or {}
                # 길이
                if 'max_len' in rule:
                    if len(s) > rule['max_len']:
                        return False, f"VR={std_vr}: value length {len(s)} exceeds max {rule['max_len']}", std or {}

        # ── Step 3: Enum 검증 ────────────────────
        enum_spec = DICOM_ENUMS.get(tag_tuple)
        if enum_spec:
            check_idx = enum_spec.get('check_index', None)
            if check_idx is not None:
                if len(values) > check_idx:
                    v_str = str(values[check_idx]).strip().upper()
                    if v_str not in enum_spec['values']:
                        return False, (
                            f"Enum violation [{enum_spec['desc']}]: "
                            f"value[{check_idx}]='{v_str}' not valid. "
                            f"Allowed: {sorted(enum_spec['values'])}"
                        ), std or {}
            else:
                v_str = str(values[0]).strip().upper() if values else ""
                if v_str not in enum_spec['values']:
                    return False, (
                        f"Enum violation [{enum_spec['desc']}]: "
                        f"'{v_str}' not valid. "
                        f"Allowed: {sorted(enum_spec['values'])}"
                    ), std or {}

        # ── Step 4: 특수 규칙 ────────────────────
        special = SPECIAL_RULES.get(tag_tuple)
        if special:
            try:
                ok, msg = special['check'](values)
                if not ok:
                    ref = special.get('ref', '')
                    return False, f"{msg} ({ref})", std or {}
            except Exception as e:
                pass

    except Exception as e:
        return False, f"Validation error: {e}", std or {}

    return True, "", std or {}


# ════════════════════════════════════════════════════
# Tag Definitions
# ════════════════════════════════════════════════════
CAT1_MANDATORY = [
    {"name": "Instance Number",           "tag": (0x0020,0x0013), "vr": "IS",  "purpose": "",               "mandatory": True},
    {"name": "Series Number",             "tag": (0x0020,0x0011), "vr": "IS",  "purpose": "Derived",        "mandatory": True},
    {"name": "Image Type",                "tag": (0x0008,0x0008), "vr": "CS",  "purpose": "(3D)",           "mandatory": True},
    {"name": "Series Description",        "tag": (0x0008,0x103E), "vr": "LO",  "purpose": "SWI",            "mandatory": True},
    {"name": "Pixel Data",                "tag": (0x7FE0,0x0010), "vr": "OB",  "purpose": "",               "mandatory": True},
    {"name": "Pixel Representation",      "tag": (0x0028,0x0103), "vr": "US",  "purpose": "",               "mandatory": True},
    {"name": "Bits Stored",               "tag": (0x0028,0x0101), "vr": "US",  "purpose": "",               "mandatory": True},
    {"name": "Rows",                      "tag": (0x0028,0x0010), "vr": "US",  "purpose": "",               "mandatory": True},
    {"name": "Columns",                   "tag": (0x0028,0x0011), "vr": "US",  "purpose": "",               "mandatory": True},
    {"name": "Pixel Spacing",             "tag": (0x0028,0x0030), "vr": "DS",  "purpose": "",               "mandatory": True},
    {"name": "Window Center",             "tag": (0x0028,0x1050), "vr": "DS",  "purpose": "MIP",            "mandatory": True},
    {"name": "Window Width",              "tag": (0x0028,0x1051), "vr": "DS",  "purpose": "MIP",            "mandatory": True},
    {"name": "Image Orientation Patient", "tag": (0x0020,0x0037), "vr": "DS",  "purpose": "post",           "mandatory": True},
    {"name": "Image Position Patient",    "tag": (0x0020,0x0032), "vr": "DS",  "purpose": "",               "mandatory": True},
    {"name": "Spacing Between Slices",    "tag": (0x0018,0x0088), "vr": "DS",  "purpose": "post",           "mandatory": True},
    {"name": "Slice Thickness",           "tag": (0x0018,0x0050), "vr": "DS",  "purpose": "post",           "mandatory": True},
    {"name": "Overlay Bits Allocated",    "tag": (0x6000,0x0100), "vr": "US",  "purpose": "",               "mandatory": False},
    {"name": "Overlay Bit Position",      "tag": (0x6000,0x0102), "vr": "US",  "purpose": "",               "mandatory": False},
    {"name": "Overlay Data",              "tag": (0x6000,0x3000), "vr": "OB",  "purpose": "",               "mandatory": False},
    {"name": "Overlay Rows",              "tag": (0x6000,0x0010), "vr": "US",  "purpose": "",               "mandatory": False},
    {"name": "Overlay Columns",           "tag": (0x6000,0x0011), "vr": "US",  "purpose": "",               "mandatory": False},
    {"name": "Diffusion b-value",         "tag": (0x0018,0x9087), "vr": "FD",  "purpose": "",               "mandatory": False},
    {"name": "Slice Location",            "tag": (0x0020,0x1041), "vr": "DS",  "purpose": "Slice Interpol", "mandatory": False},
]

CAT2_TAGS = [
    {"name": "Manufacturer",                         "tag": (0x0008,0x0070), "vr": "LO", "note": ""},
    {"name": "Number of Averages",                   "tag": (0x0018,0x0083), "vr": "DS", "note": ""},
    {"name": "Percent Sampling",                     "tag": (0x0018,0x0093), "vr": "DS", "note": ""},
    {"name": "Acquisition Matrix",                   "tag": (0x0018,0x1310), "vr": "US", "note": ""},
    {"name": "Derivation Description",               "tag": (0x0008,0x2111), "vr": "ST", "note": ""},
    {"name": "In-Plane Phase Encoding Direction",    "tag": (0x0018,0x1312), "vr": "CS", "note": ""},
    {"name": "Rows",                                 "tag": (0x0028,0x0010), "vr": "US", "note": ""},
    {"name": "Columns",                              "tag": (0x0028,0x0011), "vr": "US", "note": ""},
    {"name": "Percent Phase Field of View",          "tag": (0x0018,0x0094), "vr": "DS", "note": ""},
    {"name": "Spacing Between Slices",               "tag": (0x0018,0x0088), "vr": "DS", "note": ""},
    {"name": "Slice Thickness",                      "tag": (0x0018,0x0050), "vr": "DS", "note": ""},
    {"name": "Image Orientation Patient",            "tag": (0x0020,0x0037), "vr": "DS", "note": ""},
    {"name": "Image Position Patient",               "tag": (0x0020,0x0032), "vr": "DS", "note": ""},
    {"name": "Request Attributes Sequence",          "tag": (0x0040,0x0275), "vr": "SQ", "note": "Fonar / Other"},
    {"name": "Per-frame Functional Groups Sequence", "tag": (0x5200,0x9230), "vr": "SQ", "note": "Fonar / Other"},
    {"name": "Derivation Code Sequence",             "tag": (0x0008,0x9215), "vr": "SQ", "note": "GE / Subtraction"},
    {"name": "Field of View Dimensions",             "tag": (0x0018,0x1149), "vr": "IS", "note": "Paramed"},
]

CAT3_TAGS = [
    {"name": "Volume Based Calculation Technique",      "tag": (0x2005,0x140F), "vr": "CS", "manufacturer": "Philips",         "note": "post/[0](0008,9207)"},
    {"name": "Parallel Reduction Factor In-Plane",      "tag": (0x2005,0x140F), "vr": "FD", "manufacturer": "Philips",         "note": "[0](0018,9069)"},
    {"name": "MR Acquisition Phase Encoding Steps",     "tag": (0x2005,0x140F), "vr": "US", "manufacturer": "Philips",         "note": "[0](0018,9058)"},
    {"name": "MR Acquisition Frequency Encoding Steps", "tag": (0x2005,0x140F), "vr": "US", "manufacturer": "Philips",         "note": "[0](0018,9058)"},
    {"name": "Philips Private Creator",                 "tag": (0x2005,0x0014), "vr": "LO", "manufacturer": "Philips",         "note": ""},
    {"name": "Image Plane Number",                      "tag": (0x2001,0x100A), "vr": "IS", "manufacturer": "Philips",         "note": ""},
    {"name": "MRSeriesNrOfSlices",                      "tag": (0x2001,0x1018), "vr": "SL", "manufacturer": "Philips",         "note": ""},
    {"name": "Stack",                                   "tag": (0x2001,0x105F), "vr": "SQ", "manufacturer": "Philips",         "note": ""},
    {"name": "MRImageOffCentreAP",                      "tag": (0x2005,0x1008), "vr": "FL", "manufacturer": "Philips",         "note": ""},
    {"name": "MRImageOffCentreFH",                      "tag": (0x2005,0x1009), "vr": "FL", "manufacturer": "Philips",         "note": ""},
    {"name": "MRImageOffCentreRL",                      "tag": (0x2005,0x100A), "vr": "FL", "manufacturer": "Philips",         "note": ""},
    {"name": "SeriesDerivationDescription",             "tag": (0x2001,0x10CC), "vr": "ST", "manufacturer": "Philips",         "note": ""},
    {"name": "Siemens Private Creator",                 "tag": (0x0051,0x0010), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "Siemens Private Creator",                 "tag": (0x0021,0x0010), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "Siemens Private Creator",                 "tag": (0x0019,0x0010), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "pat factor",                              "tag": (0x0051,0x1011), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "pat factor",                              "tag": (0x0021,0x1009), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "acquisition matrix",                     "tag": (0x0051,0x100B), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "diffusion b-value",                      "tag": (0x0019,0x100C), "vr": "IS", "manufacturer": "Siemens",         "note": ""},
    {"name": "CSA HEADER1",                             "tag": (0x0029,0x1020), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "CSA HEADER2",                             "tag": (0x0021,0x1019), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "psd name1",                               "tag": (0x0019,0x109C), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "psd name2",                               "tag": (0x0019,0x109E), "vr": "LO", "manufacturer": "Siemens",         "note": ""},
    {"name": "pseq id1",                                "tag": (0x0019,0x1012), "vr": "SS", "manufacturer": "Siemens",         "note": ""},
    {"name": "pseq id2",                                "tag": (0x0025,0x1006), "vr": "SS", "manufacturer": "Siemens",         "note": ""},
    {"name": "pseq id3",                                "tag": (0x0027,0x1032), "vr": "SS", "manufacturer": "Siemens",         "note": ""},
    {"name": "slice resolution",                        "tag": (0x0019,0x1017), "vr": "DS", "manufacturer": "Siemens",         "note": ""},
    {"name": "SQ Per-frame Functional Groups Sequence", "tag": (0x5200,0x9230), "vr": "FD", "manufacturer": "Siemens",         "note": "[0](0018,9115)[0](0018,9069)"},
    {"name": "GE Private Creator",                      "tag": (0x0043,0x0010), "vr": "LO", "manufacturer": "GE",              "note": ""},
    {"name": "GE Private Creator",                      "tag": (0x0027,0x0010), "vr": "LO", "manufacturer": "GE",              "note": ""},
    {"name": "pat type",                                "tag": (0x0043,0x1084), "vr": "LO", "manufacturer": "GE",              "note": ""},
    {"name": "pat factor",                              "tag": (0x0043,0x1083), "vr": "DS", "manufacturer": "GE",              "note": ""},
    {"name": "number of frequency encoding steps",      "tag": (0x0027,0x1060), "vr": "FL", "manufacturer": "GE",              "note": ""},
    {"name": "number of phase encoding steps in-plane", "tag": (0x0027,0x1061), "vr": "FL", "manufacturer": "GE",              "note": ""},
    {"name": "Image Type (real/imaginary/phase/mag)",   "tag": (0x0043,0x102F), "vr": "SS", "manufacturer": "GE",              "note": ""},
    {"name": "Functional Protocol",                     "tag": (0x0051,0x1006), "vr": "LT", "manufacturer": "GE",              "note": "ADC source bvalue"},
    {"name": "PDB Header",                              "tag": (0x0025,0x101B), "vr": "OB", "manufacturer": "GE",              "note": ""},
    {"name": "Vas collapse flag",                       "tag": (0x0043,0x1030), "vr": "SS", "manufacturer": "GE",              "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x0029,0x1001), "vr": "SQ", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x0029,0x1002), "vr": "SQ", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x0010), "vr": "LO", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x1011), "vr": "US", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x1014), "vr": "SL", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x1016), "vr": "LO", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x1018), "vr": "SS", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "TOSHIBA_MEC",                             "tag": (0x700D,0x1019), "vr": "OB", "manufacturer": "Canon (Toshiba)", "note": ""},
    {"name": "V1",                                      "tag": (0x0011,0x1001), "vr": "OB", "manufacturer": "Esaote",          "note": ""},
    {"name": "V1",                                      "tag": (0x0011,0x1002), "vr": "DS", "manufacturer": "Esaote",          "note": ""},
    {"name": "V1",                                      "tag": (0x0011,0x1003), "vr": "DS", "manufacturer": "Esaote",          "note": ""},
    {"name": "V1",                                      "tag": (0x0011,0x1004), "vr": "DS", "manufacturer": "Esaote",          "note": ""},
    {"name": "V1",                                      "tag": (0x0011,0x1008), "vr": "DS", "manufacturer": "Esaote",          "note": ""},
    {"name": "MMCPrivate",                              "tag": (0x0029,0x102F), "vr": "",   "manufacturer": "Fonar",           "note": ""},
    {"name": "MMCPrivate",                              "tag": (0x0029,0x1032), "vr": "",   "manufacturer": "Fonar",           "note": ""},
    {"name": "MMCPrivate",                              "tag": (0x0029,0x10D7), "vr": "",   "manufacturer": "Fonar",           "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x0010), "vr": "LO", "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1001), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1002), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1003), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1004), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1005), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "Hyperfine Private Creator",               "tag": (0x351B,0x1006), "vr": "",   "manufacturer": "Hyperfine",       "note": ""},
    {"name": "acquisition voxel size",                  "tag": (0x0011,0x1017), "vr": "LO", "manufacturer": "Paramed",         "note": ""},
    {"name": "slice resolution",                        "tag": (0x0021,0x101B), "vr": "DS", "manufacturer": "Paramed",         "note": ""},
]

MANUFACTURER_KEYWORDS = {
    "Philips":         ["philips"],
    "Siemens":         ["siemens"],
    "GE":              ["ge medical", "ge healthcare", "general electric"],
    "Canon (Toshiba)": ["canon", "toshiba"],
    "Esaote":          ["esaote"],
    "Fonar":           ["fonar"],
    "Hyperfine":       ["hyperfine"],
    "Paramed":         ["paramed"],
}

# ════════════════════════════════════════════════════
# Core Functions
# ════════════════════════════════════════════════════
def get_tag_elem(ds, tag_tuple):
    try:
        tag = pydicom.tag.Tag(tag_tuple[0], tag_tuple[1])
        return ds[tag] if tag in ds else None
    except Exception:
        return None

def elem_to_display(elem):
    if elem is None:
        return None
    try:
        val = elem.value
        if val is None:
            return None
        if isinstance(val, (bytes, bytearray)):
            return f"[Binary {len(val)} bytes]"
        if isinstance(val, pydicom.sequence.Sequence):
            return f"[Sequence {len(val)} item(s)]"
        if isinstance(val, pydicom.multival.MultiValue):
            joined = ", ".join(str(v) for v in val)
            return joined[:120] + "..." if len(joined) > 120 else joined
        s = str(val)
        return s[:120] + "..." if len(s) > 120 else s
    except Exception:
        return None

def tag_to_str(tag_tuple):
    return f"({tag_tuple[0]:04X},{tag_tuple[1]:04X})"

def detect_manufacturer(ds):
    try:
        tag = pydicom.tag.Tag(0x0008, 0x0070)
        if tag in ds:
            raw = str(ds[tag].value).strip()
            if not raw:
                return None, "Unknown"
            mfr = raw.lower()
            for name, keywords in MANUFACTURER_KEYWORDS.items():
                if any(k in mfr for k in keywords):
                    return name, raw
            return None, raw
    except Exception:
        pass
    return None, "Unknown"

def is_valid_dicom(data: bytes) -> bool:
    try:
        if len(data) < 132:
            return False
        if data[128:132] == b'DICM':
            return True
        ds = pydicom.dcmread(io.BytesIO(data), force=True, stop_before_pixels=True)
        return len(ds) > 0
    except Exception:
        return False

def load_files_from_upload(uploaded_file):
    file_dict = {}
    name = uploaded_file.name.lower()
    if name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(uploaded_file.read())) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                zname    = info.filename
                basename = Path(zname).name
                if basename.startswith("__") or basename.startswith(".") or basename == "":
                    continue
                zl = zname.lower()
                if zl.endswith(".dcm") or zl.endswith(".dicom"):
                    file_dict[basename] = zf.read(zname)
                elif "." not in basename:
                    try:
                        data = zf.read(zname)
                        if is_valid_dicom(data):
                            file_dict[basename] = data
                    except Exception:
                        pass
    else:
        file_dict[uploaded_file.name] = uploaded_file.read()
    return file_dict

# ════════════════════════════════════════════════════
# Validation
# ════════════════════════════════════════════════════
def make_row(name, tag_tuple, vr_hint, extra_key, extra_val, elem, mandatory=None):
    display_val = elem_to_display(elem)
    present     = display_val is not None
    std         = get_std_info(tag_tuple)

    if present:
        is_valid, issue, _ = validate_value_full(tag_tuple, elem)
    else:
        is_valid, issue = True, ""

    # Status
    if not present:
        status = "❌  Missing" if mandatory else "⚠️  Missing"
    elif not is_valid:
        status = f"⚠️  Invalid"
    else:
        status = "✅  Present"

    is_problem = (mandatory is True) and (not present or not is_valid)

    row = {
        "Name":        name,
        "Tag":         tag_to_str(tag_tuple),
        "Std VR":      std['vr']  if std else vr_hint,
        "Std VM":      std['vm']  if std else "—",
        "Std Name":    std['name'] if std else "—",
        extra_key:     extra_val,
        "Value":       display_val if present else "MISSING",
        "Issue":       issue,
        "Status":      status,
        "_present":    present,
        "_valid":      is_valid,
        "_issue":      issue,
        "_mandatory":  mandatory,
        "_is_problem": is_problem,
    }
    return row

def validate_cat1(ds):
    results = []
    for t in CAT1_MANDATORY:
        elem = get_tag_elem(ds, t["tag"])
        row  = make_row(
            t["name"], t["tag"], t["vr"],
            "Purpose", t["purpose"],
            elem, mandatory=t["mandatory"]
        )
        row["Type"] = "🔴 Mandatory" if t["mandatory"] else "🟡 Optional"
        results.append(row)
    return results

def validate_cat2(ds):
    results = []
    for t in CAT2_TAGS:
        elem = get_tag_elem(ds, t["tag"])
        row  = make_row(
            t["name"], t["tag"], t["vr"],
            "Note", t["note"],
            elem, mandatory=False
        )
        results.append(row)
    return results

def validate_cat3(ds, mfr_name):
    if not mfr_name:
        return []
    results = []
    for t in CAT3_TAGS:
        if t["manufacturer"] != mfr_name:
            continue
        elem = get_tag_elem(ds, t["tag"])
        row  = make_row(
            t["name"], t["tag"], t["vr"],
            "Note", t["note"],
            elem, mandatory=False
        )
        row["Manufacturer"] = t["manufacturer"]
        results.append(row)
    return results

def validate_single_file(fname, file_bytes):
    try:
        ds = pydicom.dcmread(io.BytesIO(file_bytes), force=True)
    except Exception as e:
        return {"filename": fname, "error": str(e)}

    mfr_name, mfr_raw = detect_manufacturer(ds)
    cat1 = validate_cat1(ds)
    cat2 = validate_cat2(ds)
    cat3 = validate_cat3(ds, mfr_name)

    def counts(rows, mandatory_only=False):
        if mandatory_only:
            rows = [r for r in rows if r.get("_mandatory")]
        total   = len(rows)
        valid   = sum(1 for r in rows if r["_present"] and r["_valid"])
        invalid = sum(1 for r in rows if r["_present"] and not r["_valid"])
        missing = sum(1 for r in rows if not r["_present"])
        return total, valid, invalid, missing

    c1m_total, c1m_valid, c1m_invalid, c1m_missing = counts(cat1, mandatory_only=True)
    c1o_total, c1o_valid, _,            _           = counts([r for r in cat1 if not r.get("_mandatory")])
    c2_total,  c2_valid,  c2_invalid,   c2_missing  = counts(cat2)
    c3_total,  c3_valid,  c3_invalid,   c3_missing  = counts(cat3)

    cat1_problems = c1m_invalid + c1m_missing

    return {
        "filename": fname, "error": None,
        "status":   "PASS" if cat1_problems == 0 else "FAIL",
        "mfr_name": mfr_name, "mfr_raw": mfr_raw,
        "ds": ds,
        "cat1": cat1, "cat2": cat2, "cat3": cat3,
        "c1m_total": c1m_total, "c1m_valid": c1m_valid,
        "c1m_invalid": c1m_invalid, "c1m_missing": c1m_missing,
        "c1m_problems": cat1_problems,
        "c1o_total": c1o_total, "c1o_valid": c1o_valid,
        "c2_total": c2_total, "c2_valid": c2_valid,
        "c2_invalid": c2_invalid, "c2_missing": c2_missing,
        "c3_total": c3_total, "c3_valid": c3_valid,
        "c3_invalid": c3_invalid, "c3_missing": c3_missing,
    }

# ════════════════════════════════════════════════════
# Display Helpers
# ════════════════════════════════════════════════════
def to_display_df(rows, cols):
    df = pd.DataFrame(rows)
    return df[[c for c in cols if c in df.columns]]

def style_df(df):
    def row_style(row):
        s = str(row.get("Status", ""))
        if "❌" in s:
            return ["background-color: rgba(255,60,60,0.18)"] * len(row)
        elif "Invalid" in s:
            return ["background-color: rgba(255,120,0,0.18)"] * len(row)
        elif "⚠️" in s:
            return ["background-color: rgba(255,180,0,0.12)"] * len(row)
        else:
            return ["background-color: rgba(0,200,100,0.08)"] * len(row)
    return df.style.apply(row_style, axis=1)

def get_all_tags_debug(ds):
    rows = []
    for elem in ds:
        try:
            tag_tuple = (elem.tag.group, elem.tag.element)
            std       = get_std_info(tag_tuple)
            tag_str   = f"({elem.tag.group:04X},{elem.tag.element:04X})"
            if isinstance(elem.value, (bytes, bytearray)):
                val = f"[Binary {len(elem.value)} bytes]"
            elif isinstance(elem.value, pydicom.sequence.Sequence):
                val = f"[Sequence {len(elem.value)} item(s)]"
            elif hasattr(elem.value, '__iter__') and not isinstance(elem.value, str):
                val = ", ".join(str(v) for v in elem.value)[:100]
            else:
                val = str(elem.value)[:100]
            rows.append({
                "Tag":      tag_str,
                "Keyword":  elem.keyword or "—",
                "Std VR":   std['vr']  if std else elem.VR,
                "Std VM":   std['vm']  if std else "—",
                "Std Name": std['name'] if std else "—",
                "Value":    val,
            })
        except Exception:
            pass
    return pd.DataFrame(rows)

def build_export_df(result):
    rows = []
    def status_str(r, cat):
        if r["_present"] and r["_valid"]:   return "Present"
        if r["_present"] and not r["_valid"]: return "INVALID"
        return "MISSING" if (cat == 1 and r.get("_mandatory")) else "Missing"

    for r in result["cat1"]:
        rows.append({
            "File": result["filename"], "Category": "1. Mandatory-Public",
            "Type": "Mandatory" if r.get("_mandatory") else "Optional",
            "Manufacturer": "—", "Name": r["Name"], "Tag": r["Tag"],
            "Std VR": r.get("Std VR",""), "Std VM": r.get("Std VM",""),
            "Purpose/Note": r.get("Purpose",""),
            "Status": status_str(r, 1), "Value": r["Value"], "Issue": r["_issue"],
        })
    for r in result["cat2"]:
        rows.append({
            "File": result["filename"], "Category": "2. Required-Public",
            "Type": "Required", "Manufacturer": "—",
            "Name": r["Name"], "Tag": r["Tag"],
            "Std VR": r.get("Std VR",""), "Std VM": r.get("Std VM",""),
            "Purpose/Note": r.get("Note",""),
            "Status": status_str(r, 2), "Value": r["Value"], "Issue": r["_issue"],
        })
    for r in result["cat3"]:
        rows.append({
            "File": result["filename"], "Category": "3. Required-Private-MRI",
            "Type": "Required", "Manufacturer": r.get("Manufacturer",""),
            "Name": r["Name"], "Tag": r["Tag"],
            "Std VR": r.get("Std VR",""), "Std VM": r.get("Std VM",""),
            "Purpose/Note": r.get("Note",""),
            "Status": status_str(r, 3), "Value": r["Value"], "Issue": r["_issue"],
        })
    return pd.DataFrame(rows)

def excel_export(df, summary_df=None):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tag Report")
        ws = writer.sheets["Tag Report"]
        from openpyxl.styles import PatternFill
        status_col = None
        for i, cell in enumerate(ws[1]):
            if cell.value == "Status":
                status_col = i
                break
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = row[status_col].value if status_col is not None else ""
            color  = ("FFCCCC" if status == "MISSING"
                      else "FFE0CC" if status == "INVALID"
                      else "FFF3CC" if status == "Missing"
                      else "CCFFDD")
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=color)
        if summary_df is not None:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            ws2 = writer.sheets["Summary"]
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                status = row[1].value if len(row) > 1 else ""
                color  = "FFCCCC" if status == "FAIL" else "CCFFDD" if status == "PASS" else "FFE5CC"
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=color)
    return buf.getvalue()

# ════════════════════════════════════════════════════
# UI
# ════════════════════════════════════════════════════
st.markdown("""
<div class="phi-notice">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
        <span style="font-size:20px;">🔐</span>
        <span style="font-size:15px;font-weight:800;letter-spacing:1px;">Privacy & Data Handling Notice</span>
    </div>
    <div style="font-size:13px;line-height:2.0;">
        📋 This tool is intended for <b>internal QA and compatibility validation</b> purposes only.<br>
        💻 Files are processed <b>in-memory within your active session</b> — no data is explicitly saved or written to disk by this application.<br>
        🌐 Uploaded data is transmitted to and temporarily held on <b>Streamlit Cloud servers</b> (operated by Snowflake Inc.) during your session.<br>
        🔒 <b>For files containing real PHI, use only on a self-hosted / on-premise deployment</b> of this tool.<br>
        ✅ Always use <b>de-identified or anonymized DICOM files</b> when using this cloud-hosted version.
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="section-card">
  <div class="section-title">
    <div style="width:32px;height:32px;background:linear-gradient(135deg,#00d4ff,#0066ff);
        border-radius:50%;display:flex;align-items:center;justify-content:center;
        font-weight:800;color:white;font-size:15px;flex-shrink:0;">1</div>
    Upload DICOM File
  </div>
  <div style="font-size:13px;opacity:0.7;margin-top:-8px;">
    Supports single <b>.dcm</b> file or <b>.zip</b> archive containing multiple DICOM files
  </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Upload a DICOM file or ZIP archive",
    type=["dcm","DCM","zip","ZIP"],
    help="Single .dcm or .zip with multiple DICOM files"
)

if uploaded:
    with st.spinner("📂 Loading files..."):
        file_dict = load_files_from_upload(uploaded)

    if not file_dict:
        st.error("❌ No valid DICOM files found.")
        st.stop()

    total_files = len(file_dict)
    st.success(
        f"✅ **{uploaded.name}** — "
        f"{'ZIP archive' if uploaded.name.lower().endswith('.zip') else 'Single DICOM'} | "
        f"**{total_files}** file(s) detected"
    )

    with st.spinner("🔍 Validating DICOM tags (5-step standard check)..."):
        all_results = []
        prog = st.progress(0)
        for i, (fname, fbytes) in enumerate(file_dict.items()):
            all_results.append(validate_single_file(fname, fbytes))
            prog.progress((i+1)/total_files)
        prog.empty()

    valid_results = [r for r in all_results if not r.get("error")]
    error_results = [r for r in all_results if r.get("error")]
    fail_results  = [r for r in valid_results if r["status"] == "FAIL"]
    pass_results  = [r for r in valid_results if r["status"] == "PASS"]
    worst_files   = sorted(fail_results, key=lambda x: x["c1m_problems"], reverse=True)
    total_pass    = len(pass_results)
    total_fail    = len(fail_results)
    total_error   = len(error_results)

    # ── Overall Summary ───────────────────────────────
    st.markdown("""
    <div class="section-card">
      <div class="section-title">
        <div style="width:32px;height:32px;background:linear-gradient(135deg,#00d4ff,#0066ff);
            border-radius:50%;display:flex;align-items:center;justify-content:center;
            font-weight:800;color:white;font-size:15px;flex-shrink:0;">2</div>
        Overall Summary
        <span style="font-size:13px;font-weight:400;opacity:0.6;margin-left:4px;">
          — Cat.1 Mandatory: missing + invalid = FAIL
        </span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if total_fail == 0 and total_error == 0:
        oc,oi,ot = "overall-pass","✅","ALL PASS"
        os_ = f"All {total_files} file(s) passed 5-step validation. SwiftMR processing is possible."
    elif total_fail == total_files:
        oc,oi,ot = "overall-fail","❌","ALL FAIL"
        os_ = f"All {total_files} file(s) have Mandatory tag problems."
    else:
        oc,oi,ot = "overall-warning","⚠️","PARTIAL FAIL"
        os_ = f"{total_fail} of {total_files} file(s) have Mandatory tag problems."

    st.markdown(f"""
    <div class="summary-card {oc}">
        <div class="overall-title">{oi} {ot}</div>
        <div class="overall-sub">{os_}</div>
    </div>
    """, unsafe_allow_html=True)

    pass_rate  = int(total_pass/total_files*100) if total_files > 0 else 0
    rate_color = "#00c864" if pass_rate==100 else "#ffb400" if pass_rate>0 else "#ff4444"
    c1,c2,c3,c4,c5 = st.columns(5)
    for col,val,label,color in [
        (c1,str(total_files),"Total Files","#00d4ff"),
        (c2,str(total_pass), "✅ Pass",    "#00c864"),
        (c3,str(total_fail), "❌ Fail",    "#ff4444"),
        (c4,str(total_error),"⚠️ Error",   "#ffb400"),
        (c5,f"{pass_rate}%", "Pass Rate", rate_color),
    ]:
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value" style="color:{color};">{val}</div>
                <div class="metric-label">{label}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Most Problematic ─────────────────────────────
    if worst_files:
        st.markdown("""
        <div class="section-card">
          <div class="section-title">
            <div style="width:32px;height:32px;background:linear-gradient(135deg,#ff4444,#cc0000);
                border-radius:50%;display:flex;align-items:center;justify-content:center;
                font-weight:800;color:white;font-size:15px;flex-shrink:0;">3</div>
            Most Problematic Files
          </div>
        </div>
        """, unsafe_allow_html=True)

        for rank, r in enumerate(worst_files[:10], 1):
            prob_tags  = [x for x in r["cat1"] if x["_is_problem"]]
            prob_names = " · ".join([
                f'<b>{x["Name"]}</b>'
                + (f' <span style="font-size:10px;color:#ff8c00;">[{x["_issue"][:40]}]</span>'
                   if x["_present"] and not x["_valid"] else "")
                for x in prob_tags
            ])
            cat3_label = f"{r['c3_valid']}/{r['c3_total']} ({r['mfr_name']})" if r["mfr_name"] else "N/A"
            st.markdown(f"""
            <div class="file-problem-card">
                <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                    <span style="font-size:18px;font-weight:900;color:#ff4444;">#{rank}</span>
                    <span style="font-size:14px;font-weight:700;">{r['filename']}</span>
                    <span style="margin-left:auto;font-size:12px;background:rgba(255,60,60,0.2);
                        color:#ff4444;padding:2px 10px;border-radius:20px;font-weight:700;">
                        ❌ {r['c1m_problems']} Problem(s)
                    </span>
                </div>
                <div style="font-size:12px;opacity:0.8;line-height:1.8;">
                    🏭 <b>{r['mfr_raw']}</b>
                    &nbsp;·&nbsp; Cat.1 Mandatory: {r['c1m_valid']}/{r['c1m_total']} valid
                    &nbsp;·&nbsp; Invalid: {r['c1m_invalid']} &nbsp;·&nbsp; Missing: {r['c1m_missing']}
                    &nbsp;·&nbsp; Cat.2: {r['c2_valid']}/{r['c2_total']}
                    &nbsp;·&nbsp; Cat.3: {cat3_label}
                </div>
                <div style="font-size:12px;margin-top:6px;color:#ff6666;">Problems: {prob_names}</div>
            </div>
            """, unsafe_allow_html=True)

    for r in error_results:
        st.markdown(f"""
        <div class="file-problem-card">
            <div style="display:flex;align-items:center;gap:10px;">
                <span style="font-size:14px;font-weight:700;">{r['filename']}</span>
                <span style="margin-left:auto;font-size:12px;background:rgba(255,60,60,0.2);
                    color:#ff4444;padding:2px 10px;border-radius:20px;font-weight:700;">⚠️ Read Error</span>
            </div>
            <div style="font-size:12px;opacity:0.7;margin-top:4px;">{r['error']}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── File Detail ───────────────────────────────────
    st.markdown("""
    <div class="section-card">
      <div class="section-title">
        <div style="width:32px;height:32px;background:linear-gradient(135deg,#00d4ff,#0066ff);
            border-radius:50%;display:flex;align-items:center;justify-content:center;
            font-weight:800;color:white;font-size:15px;flex-shrink:0;">4</div>
        File-by-File Detail
      </div>
    </div>
    """, unsafe_allow_html=True)

    def make_label(r):
        if r.get("error"):
            return f"⚠️  {r['filename']}  [ERROR]"
        icon = "✅" if r["status"] == "PASS" else "❌"
        prob = r.get("c1m_problems", 0)
        suf  = f"  — {prob} problem(s)" if prob > 0 else ""
        return f"{icon}  {r['filename']}{suf}"

    opts = [make_label(r) for r in all_results]
    def_idx = 0
    if worst_files:
        wf = worst_files[0]["filename"]
        for i, r in enumerate(all_results):
            if r["filename"] == wf:
                def_idx = i; break

    sel_label  = st.selectbox("Select file", options=opts, index=def_idx, key="file_sel")
    sel_result = all_results[opts.index(sel_label)]
    st.markdown("<br>", unsafe_allow_html=True)

    if sel_result.get("error"):
        st.error(f"❌ Cannot read: **{sel_result['filename']}**\n\n{sel_result['error']}")
    else:
        r = sel_result

        # 상태 배너
        if r["status"] == "PASS":
            bc,bi,bt = "overall-pass","✅","PASS"
            bs = "All Mandatory-Public tags are present and valid. SwiftMR processing is possible."
        else:
            bc,bi,bt = "overall-fail","❌","FAIL"
            bs = f"Cat.1 Mandatory — Missing: {r['c1m_missing']}, Invalid: {r['c1m_invalid']}. SwiftMR cannot process."

        st.markdown(f"""
        <div class="summary-card {bc}" style="padding:16px 20px;margin-bottom:16px;">
            <div style="font-size:20px;font-weight:900;letter-spacing:1px;margin-bottom:2px;">
                {bi} {bt} — {r['filename']}
            </div>
            <div style="font-size:13px;opacity:0.8;">{bs}</div>
        </div>
        """, unsafe_allow_html=True)

        # Manufacturer
        mfr_display = r["mfr_name"] if r["mfr_name"] else "Not Detected"
        mfr_color   = "#00d4ff" if r["mfr_name"] else "#ffb400"
        st.markdown(f"""
        <div style="margin-bottom:16px;">
            <span class="manufacturer-badge" style="color:{mfr_color};">
                🏭 Manufacturer: <b>{mfr_display}</b>
                &nbsp;·&nbsp; Raw: <i>{r['mfr_raw']}</i>
            </span>
        </div>
        """, unsafe_allow_html=True)

        # DICOM Standard 정보 박스
        st.markdown("""
        <div class="std-info-box">
            <b>🔬 Validation Engine v2.0</b> — 5-step DICOM Standard check:<br>
            &nbsp;&nbsp;① VM count (pydicom DicomDictionary)
            &nbsp;&nbsp;② VR type/range/pattern (PS3.5)
            &nbsp;&nbsp;③ Enum values (PS3.3)
            &nbsp;&nbsp;④ Tag-specific rules (PS3.3 references)
            &nbsp;&nbsp;⑤ Placeholder/zero detection
        </div>
        """, unsafe_allow_html=True)

        # Debug
        with st.expander("🔍 Debug: All Tags in This DICOM File", expanded=False):
            debug_df = get_all_tags_debug(r["ds"])
            st.dataframe(debug_df, use_container_width=True, hide_index=True, height=400)
            st.caption(f"Total tags: **{len(debug_df)}** · Std VR/VM from pydicom DicomDictionary")

        # ── Cat.1 ─────────────────────────────────────
        mand_color = "#00c864" if r["c1m_problems"] == 0 else "#ff4444"
        st.markdown(f"""
        <div class="cat1-header">
            <div style="font-size:17px;font-weight:800;margin-bottom:6px;">
                🔴 Category 1 — Mandatory-Public
            </div>
            <div style="font-size:13px;opacity:0.85;line-height:1.9;">
                Valid: <span style="color:{mand_color};font-weight:700;">{r['c1m_valid']}/{r['c1m_total']}</span>
                &nbsp;·&nbsp; <span style="color:#ff8c00;font-weight:600;">Invalid: {r['c1m_invalid']}</span>
                &nbsp;·&nbsp; <span style="color:#ff4444;font-weight:600;">Missing: {r['c1m_missing']}</span>
                &nbsp;·&nbsp; Optional: <span style="color:#ffb400;">{r['c1o_valid']}/{r['c1o_total']}</span>
                &nbsp;·&nbsp; <span style="opacity:0.6;font-size:12px;">PASS = all mandatory valid & present</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        prob_tags = [x for x in r["cat1"] if x["_is_problem"]]
        if prob_tags:
            items = "".join([
                f'<div style="font-size:13px;margin:5px 0;display:flex;align-items:flex-start;gap:8px;">'
                f'<span>{"❌" if not x["_present"] else "⚠️"}</span>'
                f'<div><b>{x["Name"]}</b> '
                f'<span style="font-family:monospace;font-size:11px;opacity:0.5;">{x["Tag"]}</span>'
                f'<span style="opacity:0.6;font-size:11px;"> · Std VR={x.get("Std VR","?")} VM={x.get("Std VM","?")}</span>'
                + (f'<br><span style="color:#ff8c00;font-size:12px;">↳ {x["_issue"]}</span>' if x["_issue"] else "")
                + "</div></div>"
                for x in prob_tags
            ])
            st.markdown(f"""
            <div style="background:rgba(255,60,60,0.1);border:1.5px solid rgba(255,60,60,0.4);
                border-left:4px solid #ff4444;border-radius:12px;padding:14px 20px;margin-bottom:12px;">
                <div style="font-weight:800;color:#ff4444;margin-bottom:8px;font-size:14px;">
                    ❌ Mandatory Tag Problems — SwiftMR Cannot Process
                </div>{items}
            </div>
            """, unsafe_allow_html=True)

        df1 = to_display_df(r["cat1"], ["Name","Tag","Std VR","Std VM","Purpose","Type","Value","Issue","Status"])
        st.dataframe(style_df(df1), use_container_width=True, hide_index=True,
                     height=min(50+len(df1)*35, 700))

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Cat.2 ─────────────────────────────────────
        c2_color = "#00c864" if (r["c2_invalid"]+r["c2_missing"])==0 else "#ffb400"
        st.markdown(f"""
        <div class="cat2-header">
            <div style="font-size:17px;font-weight:800;margin-bottom:6px;">
                🟠 Category 2 — Required-Public
            </div>
            <div style="font-size:13px;opacity:0.85;line-height:1.9;">
                Valid: <span style="color:{c2_color};font-weight:700;">{r['c2_valid']}/{r['c2_total']}</span>
                &nbsp;·&nbsp; <span style="color:#ff8c00;">Invalid: {r['c2_invalid']}</span>
                &nbsp;·&nbsp; <span style="color:#ffb400;">Missing: {r['c2_missing']}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        df2 = to_display_df(r["cat2"], ["Name","Tag","Std VR","Std VM","Note","Value","Issue","Status"])
        st.dataframe(style_df(df2), use_container_width=True, hide_index=True,
                     height=min(50+len(df2)*35, 650))

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Cat.3 ─────────────────────────────────────
        if r["mfr_name"]:
            c3_color = "#00c864" if (r["c3_invalid"]+r["c3_missing"])==0 else "#ffb400"
            st.markdown(f"""
            <div class="cat3-header">
                <div style="font-size:17px;font-weight:800;margin-bottom:6px;">
                    🔵 Category 3 — Required-Private-MRI
                </div>
                <div style="font-size:13px;opacity:0.85;line-height:1.9;">
                    Manufacturer: <b>{r['mfr_name']}</b>
                    &nbsp;·&nbsp; Valid: <span style="color:{c3_color};font-weight:700;">{r['c3_valid']}/{r['c3_total']}</span>
                    &nbsp;·&nbsp; <span style="color:#ff8c00;">Invalid: {r['c3_invalid']}</span>
                    &nbsp;·&nbsp; <span style="color:#ffb400;">Missing: {r['c3_missing']}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
            df3 = to_display_df(r["cat3"], ["Manufacturer","Name","Tag","Std VR","Std VM","Note","Value","Issue","Status"])
            st.dataframe(style_df(df3), use_container_width=True, hide_index=True,
                         height=min(50+len(df3)*35, 700))
            st.caption(f"💡 Showing private tags for detected manufacturer: **{r['mfr_name']}**")
        else:
            st.markdown(f"""
            <div class="no-mfr-box">
                <div style="font-size:17px;font-weight:800;margin-bottom:8px;">
                    🔵 Category 3 — Required-Private-MRI
                </div>
                <div style="font-size:13px;line-height:1.9;">
                    ⚠️ <b>Manufacturer not detected</b> — Category 3 validation skipped.<br>
                    Raw Manufacturer value: <b>{r['mfr_raw']}</b><br>
                    <span style="opacity:0.7;font-size:12px;">
                        Supported: Philips · Siemens · GE · Canon (Toshiba) · Esaote · Fonar · Hyperfine · Paramed
                    </span>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

    # ── Export ────────────────────────────────────────
    st.markdown("""
    <div class="section-card">
      <div class="section-title">
        <div style="width:32px;height:32px;background:linear-gradient(135deg,#00d4ff,#0066ff);
            border-radius:50%;display:flex;align-items:center;justify-content:center;
            font-weight:800;color:white;font-size:15px;flex-shrink:0;">5</div>
        Export Report
      </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["📄 Current File", "📦 All Files"])

    with tab1:
        if not sel_result.get("error"):
            df_exp = build_export_df(sel_result)
            base   = sel_result["filename"].replace(".dcm","")
            c1,c2  = st.columns(2)
            with c1:
                st.download_button("⬇️ Download CSV",
                    data=df_exp.to_csv(index=False).encode("utf-8"),
                    file_name=f"tag_report_{base}.csv", mime="text/csv",
                    use_container_width=True)
            with c2:
                st.download_button("⬇️ Download Excel",
                    data=excel_export(df_exp),
                    file_name=f"tag_report_{base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with st.expander("📊 Preview", expanded=False):
                st.dataframe(df_exp, use_container_width=True, hide_index=True, height=400)

    with tab2:
        vfe = [r for r in all_results if not r.get("error")]
        if vfe:
            df_all    = pd.concat([build_export_df(r) for r in vfe], ignore_index=True)
            base_name = uploaded.name.replace(".zip","").replace(".dcm","")
            summary_rows = []
            for res in all_results:
                if res.get("error"):
                    summary_rows.append({
                        "Filename": res["filename"], "Status": "ERROR",
                        "Manufacturer": "—",
                        "Cat1 Valid":"—","Cat1 Invalid":"—","Cat1 Missing":"—",
                        "Cat2 Valid":"—","Cat2 Invalid":"—","Cat2 Missing":"—",
                        "Cat3 Valid":"—","Cat3 Invalid":"—","Cat3 Missing":"—",
                        "Error": res["error"],
                    })
                else:
                    summary_rows.append({
                        "Filename": res["filename"], "Status": res["status"],
                        "Manufacturer": res["mfr_raw"],
                        "Cat1 Valid": res["c1m_valid"],
                        "Cat1 Invalid": res["c1m_invalid"],
                        "Cat1 Missing": res["c1m_missing"],
                        "Cat2 Valid": res["c2_valid"],
                        "Cat2 Invalid": res["c2_invalid"],
                        "Cat2 Missing": res["c2_missing"],
                        "Cat3 Valid": res["c3_valid"],
                        "Cat3 Invalid": res["c3_invalid"],
                        "Cat3 Missing": res["c3_missing"],
                        "Error": "",
                    })
            df_sum = pd.DataFrame(summary_rows)
            c1,c2  = st.columns(2)
            with c1:
                st.download_button("⬇️ Download All CSV",
                    data=df_all.to_csv(index=False).encode("utf-8"),
                    file_name=f"tag_report_ALL_{base_name}.csv", mime="text/csv",
                    use_container_width=True)
            with c2:
                st.download_button("⬇️ Download All Excel",
                    data=excel_export(df_all, df_sum),
                    file_name=f"tag_report_ALL_{base_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with st.expander("📊 Preview All", expanded=False):
                st.dataframe(df_all, use_container_width=True, hide_index=True, height=400)

# ── Sidebar ──────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:16px 0 20px;">
        <div style="width:56px;height:56px;margin:0 auto 10px;
            display:flex;align-items:center;justify-content:center;">{sidebar_logo_html}</div>
        <div style="font-size:14px;font-weight:800;letter-spacing:2px;
            background:linear-gradient(90deg,#00d4ff,#0066ff);
            -webkit-background-clip:text;-webkit-text-fill-color:transparent;">SwiftMR</div>
        <div style="font-size:11px;color:#8892a4;margin-top:2px;letter-spacing:1px;">
            DICOM Tag Validator v2.0</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">🔬 Validation Engine</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:12px;line-height:2.0;">
        ① <b>VM</b> — DicomDictionary count check<br>
        ② <b>VR</b> — type / range / pattern (PS3.5)<br>
        ③ <b>Enum</b> — allowed values (PS3.3)<br>
        ④ <b>Special</b> — tag-specific rules (PS3.3)<br>
        ⑤ <b>Placeholder</b> — zero / empty detection
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">📋 Categories</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:13px;line-height:2.0;">
        <div style="margin-bottom:8px;">
            <span style="color:#ff4444;font-weight:700;">🔴 Cat.1 Mandatory-Public</span><br>
            <span style="font-size:11px;opacity:0.7;">PASS/FAIL · missing + invalid = problem</span>
        </div>
        <div style="margin-bottom:8px;">
            <span style="color:#ff8c00;font-weight:700;">🟠 Cat.2 Required-Public</span><br>
            <span style="font-size:11px;opacity:0.7;">Standard public tags · value validated</span>
        </div>
        <div>
            <span style="color:#00d4ff;font-weight:700;">🔵 Cat.3 Required-Private-MRI</span><br>
            <span style="font-size:11px;opacity:0.7;">Detected manufacturer only · skipped if unknown</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">📊 Status Legend</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:13px;line-height:2.0;">
        <span style="color:#00c864;font-weight:700;">✅ Present</span> — exists + all checks pass<br>
        <span style="color:#ff8c00;font-weight:700;">⚠️ Invalid</span> — exists but value abnormal<br>
        <span style="color:#ff4444;font-weight:700;">❌ Missing</span> — mandatory tag absent<br>
        <span style="color:#ffb400;font-weight:700;">⚠️ Missing</span> — optional tag absent
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">🔍 Validated Rules</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:11px;line-height:1.9;opacity:0.85;">
        ImagePositionPatient → VM=3, non-zero<br>
        ImageOrientationPatient → VM=6, non-zero<br>
        PixelSpacing → VM=2, positive<br>
        WindowWidth → must be &gt; 0 (PS3.3)<br>
        BitsStored → 1–16 (PS3.5)<br>
        PatientPosition → 16 standard values<br>
        Modality → DICOM standard list<br>
        ImageType[0] → ORIGINAL or DERIVED<br>
        Rows/Columns → positive integer<br>
        VR types → PS3.5 Table 6.2-1
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">🏭 Supported Manufacturers</div>', unsafe_allow_html=True)
    for mfr in MANUFACTURER_KEYWORDS.keys():
        st.markdown(f'<div style="font-size:13px;padding:2px 0;">· {mfr}</div>', unsafe_allow_html=True)

    st.divider()
    st.markdown('<div class="sidebar-section-title">🔐 Data Privacy</div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-size:12px;line-height:1.9;opacity:0.85;">
        💻 In-memory processing only<br>
        🌐 Streamlit Cloud (Snowflake) servers<br>
        🔒 PHI → use self-hosted deployment<br>
        ✅ De-identified files recommended
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown("""
    <div style="text-align:center;font-size:11px;color:#4a5568;padding:8px 0;">
        © 2024 AIRS Medical Inc.<br>All rights reserved.
    </div>
    """, unsafe_allow_html=True)
